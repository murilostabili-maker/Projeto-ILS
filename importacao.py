"""
importacao.py
--------------
"""

import io
import time
from typing import List, Tuple

import pandas as pd

import osm
from dados import Cliente

COLUNAS_TEMPLATE = ["nome", "endereco", "latitude", "longitude", "demanda"]


def gerar_template_bytes() -> bytes:
    """Gera, em memória, uma planilha .xlsx modelo para o usuário preencher e enviar de volta."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    fonte_padrao = Font(name="Arial", size=11)
    fonte_cabecalho = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    preenchimento_cabecalho = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    preenchimento_instrucao = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    instrucoes = (
        "Preencha uma linha por cliente abaixo do cabeçalho (linha 4). "
        "'nome' e 'demanda' são obrigatórios. Para localizar o cliente, preencha ou "
        "'endereco' ou 'latitude' + 'longitude' (se latitude/longitude estiverem "
        "preenchidas, o endereço é ignorado). "
        "Apague as linhas de exemplo antes de importar."
    )
    ws.merge_cells("A1:E3")
    celula_instrucao = ws["A1"]
    celula_instrucao.value = instrucoes
    celula_instrucao.font = fonte_padrao
    celula_instrucao.alignment = Alignment(wrap_text=True, vertical="top")
    celula_instrucao.fill = preenchimento_instrucao

    linha_cabecalho = 4
    for col_idx, nome_coluna in enumerate(COLUNAS_TEMPLATE, start=1):
        c = ws.cell(row=linha_cabecalho, column=col_idx, value=nome_coluna)
        c.font = fonte_cabecalho
        c.fill = preenchimento_cabecalho

    exemplos = [
        ["Cliente Exemplo 1 (por endereço)", "Rua das Trincheiras, João Pessoa, PB", "", "", 150],
        ["Cliente Exemplo 2 (por coordenadas)", "", -7.1139, -34.8639, 200],
    ]
    for i, linha in enumerate(exemplos, start=linha_cabecalho + 1):
        for col_idx, valor in enumerate(linha, start=1):
            c = ws.cell(row=i, column=col_idx, value=valor)
            c.font = fonte_padrao

    larguras = {"A": 32, "B": 40, "C": 14, "D": 14, "E": 10}
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura

    ws.freeze_panes = ws.cell(row=linha_cabecalho + 1, column=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def ler_planilha(arquivo) -> pd.DataFrame:
    """
    Lê o arquivo enviado (.csv ou .xlsx/.xls) e devolve um DataFrame já com as
    colunas de COLUNAS_TEMPLATE. Encontra a linha de cabeçalho automaticamente
    (funciona tanto com a planilha modelo, que tem o cabeçalho na linha 4,
    quanto com um CSV simples cujo cabeçalho é a primeira linha).
    """
    nome_arquivo = arquivo.name.lower()

    if nome_arquivo.endswith(".csv"):
        df_bruto = pd.read_csv(arquivo, header=None, dtype=str)
    elif nome_arquivo.endswith((".xlsx", ".xls")):
        df_bruto = pd.read_excel(arquivo, sheet_name=0, header=None)
    else:
        raise ValueError("Formato não suportado. Envie um arquivo .csv ou .xlsx.")

    linha_cabecalho = None
    for i, linha in df_bruto.iterrows():
        valores = [str(v).strip().lower() for v in linha.values]
        if "nome" in valores and "demanda" in valores:
            linha_cabecalho = i
            break

    if linha_cabecalho is None:
        raise ValueError(
            "Não encontrei a linha de cabeçalho (nome, endereco, latitude, longitude, "
            "demanda) na planilha. Use a planilha modelo como base."
        )

    df = df_bruto.iloc[linha_cabecalho + 1:].reset_index(drop=True)
    df.columns = [str(v).strip().lower() for v in df_bruto.iloc[linha_cabecalho].values]

    faltando = [c for c in COLUNAS_TEMPLATE if c not in df.columns]
    if faltando:
        raise ValueError(f"Colunas faltando na planilha: {', '.join(faltando)}")

    return df


def processar_linhas(df: pd.DataFrame, id_inicial: int, pausa_geocoding: float = 1.0) -> Tuple[List[Cliente], List[str]]:
    """
    Valida e converte cada linha do DataFrame em um Cliente.

    Geocodifica via Nominatim apenas quando latitude/longitude não vierem
    preenchidas na planilha (para economizar chamadas e respeitar o limite de
    uso do serviço público — por isso uma pequena pausa entre geocodificações).

    Retorna (clientes_validos, mensagens_de_erro). Linhas com problema são
    ignoradas (não interrompem o processamento das demais).
    """
    clientes: List[Cliente] = []
    erros: List[str] = []
    proximo_id = id_inicial

    for idx, linha in df.iterrows():
        num_linha_planilha = idx + 2  # referência aproximada para o usuário localizar a linha

        nome = str(linha.get("nome", "")).strip()
        if not nome or nome.lower() == "nan":
            erros.append(f"Linha {num_linha_planilha}: 'nome' vazio — linha ignorada.")
            continue

        demanda_bruta = linha.get("demanda", None)
        try:
            demanda = int(float(demanda_bruta))
            if demanda < 0:
                raise ValueError
        except (TypeError, ValueError):
            erros.append(f"Linha {num_linha_planilha} ({nome}): 'demanda' inválida — linha ignorada.")
            continue

        lat_bruta = linha.get("latitude", None)
        lon_bruta = linha.get("longitude", None)
        latitude = longitude = None
        tem_coords = False
        try:
            if (
                pd.notna(lat_bruta) and pd.notna(lon_bruta)
                and str(lat_bruta).strip() != "" and str(lon_bruta).strip() != ""
            ):
                latitude = float(lat_bruta)
                longitude = float(lon_bruta)
                tem_coords = True
        except (TypeError, ValueError):
            erros.append(f"Linha {num_linha_planilha} ({nome}): latitude/longitude inválidas — tentando por endereço.")

        if not tem_coords:
            endereco = str(linha.get("endereco", "")).strip()
            if not endereco or endereco.lower() == "nan":
                erros.append(f"Linha {num_linha_planilha} ({nome}): sem endereço nem latitude/longitude — linha ignorada.")
                continue
            try:
                resultados = osm.buscar_endereco(endereco, limite=1)
                if not resultados:
                    erros.append(f"Linha {num_linha_planilha} ({nome}): endereço '{endereco}' não encontrado — linha ignorada.")
                    continue
                latitude = float(resultados[0]["lat"])
                longitude = float(resultados[0]["lon"])
            except RuntimeError as e:
                erros.append(f"Linha {num_linha_planilha} ({nome}): erro ao geocodificar — {e}")
                continue
            finally:
                time.sleep(pausa_geocoding)  # respeita o limite de uso do Nominatim (~1 req/s)

        clientes.append(Cliente(id=proximo_id, nome=nome, latitude=latitude, longitude=longitude, demanda=demanda))
        proximo_id += 1

    return clientes, erros
